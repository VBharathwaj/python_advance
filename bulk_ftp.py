import os
from configparser import SafeConfigParser
import xlsxwriter
import openpyxl
import fnmatch

cur_wrk_dir = os.getcwd()
submitted_by=''
acquisition_source=''
ticket_number=''
deal_id=''
loan_mapping_file_path = ''
parser = SafeConfigParser()
parser.read(cur_wrk_dir+"\\Inputs\\config.ini")
ingestion_path = (parser.get('paths', 'ingestion_path'))
source_path = (parser.get('paths', 'source_path'))
row = 0
col = 0
loan_number_mapping = []
dropped_loan_mapping = []
cross_reference_sheet = ''
dropped_loans_sheet = ''
final_data = []
doc_type_map=[]

#Step 2 - Reading the Config FIle
def fetch_input():
	global ticket_number
	global acquisition_source
	global deal_id
	global submitted_by
	submitted_by = (parser.get('user_inputs', 'submitted_by'))
	acquisition_source = (parser.get('user_inputs', 'acquisition_source'))
	ticket_number = (parser.get('user_inputs', 'ticket_number'))
	deal_id = (parser.get('user_inputs', 'deal_id'))

#Step 3 - Get verification from user
def verify_input():
	print('Verify Inputs\n')
	print('Submitted By 		: ' + submitted_by)
	print('Acquisition Source	: ' + acquisition_source)
	print('Ticket Number 		: ' + ticket_number)
	print('Deal ID 		: ' + deal_id)
	a = input('\nProcees?\n1)Yes\n2)No...')
	return a
		
#Step 5 - Fetch the file names from the directory
def fetch_file_names():
	global loan_mapping_file_path
	image_path = ingestion_path+ticket_number+'\\'
	for dirpath,_,filenames in os.walk(image_path):
		for f in filenames:
			final_file_path = os.path.abspath(os.path.join(dirpath, f))
			if "cross reference" in final_file_path.lower():
				loan_mapping_file_path = final_file_path
			if "report" in final_file_path.lower() or ".csv" in final_file_path.lower() or "thumbs.db" in final_file_path.lower() or ".xlsx" in final_file_path.lower():
				continue
			temp={}
			temp['File Path/Name']=final_file_path
			temp['Prev Doc Type']=''
			temp['NSM Doc Type']=''
			temp['Prev #']=''
			temp['NSM #']=''
			temp['File Name']=''
			temp['Submitted By']=submitted_by
			temp['Acquisition Source']=acquisition_source
			temp['Ticket#']=ticket_number
			temp['DealID']=deal_id
			temp['Doc Date']=''
			final_data.append(temp)
			del(temp)
	
#Extract Data
def extract():
	for single_record in final_data:
		single_record_split=single_record['File Path/Name'].split("\\")
		temp_filename = single_record_split[len(single_record_split)-1]
		single_record['File Name'] = temp_filename
		temp_filename_split = temp_filename.split("_")
		single_record['Prev Doc Type'] = temp_filename_split[2]
		single_record['Prev #'] = temp_filename_split[0]
		temp_date = temp_filename_split[3].split(".")[0]
		temp_formatted_date=temp_date[4:6] + "\\" + temp_date[6:] + "\\" + temp_date[0:4]
		single_record['Doc Date'] = temp_formatted_date
	
#Copy doc type mappings from mappings excel
def read_doc_type_mappings():
	doc_type_mappings_excel_path = (parser.get('paths', 'doc_type_mapping_sheet_path'))
	wb = openpyxl.load_workbook(doc_type_mappings_excel_path)
	sheet = wb.get_sheet_by_name('Sheet1')
	for row in sheet:
		temp ={}
		temp['Prev'] = row[0].value
		temp['Cur'] = row[1].value
		doc_type_map.append(temp)
		del(temp)
	wb.close()
	
#Mapping the current doc type
def map_current_doc_type():
	for single_record in final_data:
		temp = single_record['Prev Doc Type']
		for single_map in doc_type_map:
			if temp.lower() == str(single_map['Prev']).lower():
				single_record['NSM Doc Type'] = single_map['Cur']
	
	for dict in final_data:
		if str(dict['NSM Doc Type']).lower() == '':
			dict['NSM Doc Type'] = 'NA'
	
def check_final_cross_reference(sheet_names):
	global cross_reference_sheet
	for sheet in sheet_names:
		if "final cross reference" in sheet.lower():
			cross_reference_sheet = sheet
			return True
			
def check_dropped_loan_flag(sheet_names):
	global dropped_loans_sheet
	for sheet in sheet_names:
		if "dropped" in sheet.lower():
			dropped_loans_sheet = sheet
			return True
	
def process_cross_reference_worksheet(wb):
	cross_reference_worksheet = wb.get_sheet_by_name(cross_reference_sheet)
	servicer_loan_column_ref = ''
	msn_loan_column_ref = ''
	for row in cross_reference_worksheet:
		for cell in row:
			if "servicer loan" in str(cell.value).lower():
				cell_ref = str(cell)
				servicer_loan_column_ref = ord(cell_ref[len(cell_ref)-3])-65
			if "msn loan" in str(cell.value).lower():
				cell_ref = str(cell)
				msn_loan_column_ref = ord(cell_ref[len(cell_ref)-3])-65
				
	for row in cross_reference_worksheet:			
		temp ={}
		if "servicer loan" in str(row[servicer_loan_column_ref].value).lower():
			continue
		temp['Servicer'] = row[servicer_loan_column_ref].value
		temp['MSN'] = row[msn_loan_column_ref].value
		loan_number_mapping.append(temp)
		del(temp)
	
def process_dropped_loans_worksheet(wb):
	dropped_loans_worksheet = wb.get_sheet_by_name(dropped_loans_sheet)
	servicer_loan_column_ref = ''
	for row in dropped_loans_worksheet:
		for cell in row:
			if "servicer loan" in str(cell.value).lower():
				cell_ref = str(cell)
				servicer_loan_column_ref = ord(cell_ref[len(cell_ref)-3])-65
				
	for row in dropped_loans_worksheet:			
		temp ={}
		if "servicer loan" in str(row[servicer_loan_column_ref].value).lower():
			continue
		temp['Servicer'] = row[servicer_loan_column_ref].value
		temp['MSN'] = 'Dropped'
		loan_number_mapping.append(temp)
		del(temp)
	
def process_with_dropped(wb):
	process_cross_reference_worksheet(wb)
	process_dropped_loans_worksheet(wb)

#Copy the previous and current Loan Number from Cross  Reference
def read_loan_mappings():
	global loan_mapping_file_path
	wb = openpyxl.load_workbook(loan_mapping_file_path)
	sheet_names = wb.get_sheet_names()
	final_cross_reference_flag = check_final_cross_reference(sheet_names)
	dropped_loan_flag = check_dropped_loan_flag(sheet_names)
	if final_cross_reference_flag and dropped_loan_flag:
		process_with_dropped(wb)
	elif final_cross_reference_flag and not(dropped_loan_flag):
		print("process_without_dropped")
	else:
		print("process_without_final_cross_ref")
	wb.close()
				
def map_current_loan():
	for single_record in final_data:
		temp = single_record['Prev #']
		for single_map in loan_number_mapping:
			if temp.lower() == str(single_map['Servicer']).lower():
				single_record['NSM #'] = single_map['MSN']
	
#Insert Excel Header
def excel_header():
	global worksheet
	global workbook
	global row
	global col
	worksheet.write(row,col,"File Path/Name")
	col+=1
	worksheet.write(row,col,"Prev Doc Type")
	col+=1
	worksheet.write(row,col,"NSM Doc Type")
	col+=1
	worksheet.write(row,col,"Prev #")
	col+=1
	worksheet.write(row,col,"NSM #")
	col+=1
	worksheet.write(row,col,"File Name")
	col+=1
	worksheet.write(row,col,"Submitted By")
	col+=1
	worksheet.write(row,col,"Acquisition Source")
	col+=1
	worksheet.write(row,col,"Ticket#")
	col+=1
	worksheet.write(row,col,"DealID")
	col+=1
	worksheet.write(row,col,"Doc Date")
	row+=1
	col=0
				
#Write status to excel
def write_to_excel():
	global worksheet
	global na_worksheet
	global dropped_worksheet
	global workbook
	global row
	global col
	
	excel_header()

	for dict in final_data:
		if str(dict['NSM #']).lower() == 'dropped' or str(dict['NSM Doc Type']) == 'NA':
			continue
		for key in dict.keys():
			worksheet.write(row,col,str(dict[key]))
			col+=1
		row+=1
		col=0
			
def write_dropped_to_excel():
	global worksheet
	global dropped_worksheet
	global workbook
	row_dropped=0
	col_dropped=0
	for dict in final_data:
		if str(dict['NSM #']).lower() == 'dropped':
			for key in dict.keys():
				dropped_worksheet.write(row_dropped,col_dropped,str(dict[key]))
				col_dropped+=1
			row_dropped+=1
			col_dropped=0
	
def write_na_to_excel():
	global worksheet
	global na_worksheet
	global workbook
	row_na=0
	col_na=0
	for dict in final_data:
		if str(dict['NSM Doc Type']) == 'NA':
			for key in dict.keys():
				na_worksheet.write(row_na,col_na,str(dict[key]))
				col_na+=1
			row_na+=1
			col_na=0
			
def write_dt_to_excel():
	global doc_type_worksheet
	global workbook
	row = 0  
	col = 0
	
	for dict in doc_type_map:
		for key in dict.keys():
			doc_type_worksheet.write(row,col,str(dict[key]))
			col+=1
		row+=1
		col=0

def write_ln_to_excel():
	global ln_worksheet
	global workbook
	row = 0  
	col = 0
	ln_worksheet.write(row,col,'Servicer #')
	col+=1
	ln_worksheet.write(row,col,'NSM #')
	col=0
	row+=1
	
	for dict in loan_number_mapping:
		for key in dict.keys():
			ln_worksheet.write(row,col,str(dict[key]))
			col+=1
		row+=1
		col=0
	
def excel_jobs():
	write_to_excel()
	write_dt_to_excel()
	write_dropped_to_excel()
	write_na_to_excel()
	write_ln_to_excel()
	
def mappin_jobs():
	read_doc_type_mappings()
	map_current_doc_type()
	read_loan_mappings()
	map_current_loan()
	
#Step 1 - Fetch necessary input from the user
fetch_input()

#Step 3 - Verifying Input
proceed = verify_input()

if proceed == "1":
	print("\nProcessing....")
	#Step 4 - Fetch File Names
	fetch_file_names()
	#Extract data from file path
	extract()
	#Write to Excel
	output_file_name=cur_wrk_dir+"\\Output\\"+ticket_number+".xlsx"
	mappin_jobs()
	workbook = xlsxwriter.Workbook(output_file_name)
	worksheet = workbook.add_worksheet()
	worksheet.set_column(1, 3, 30)
	doc_type_worksheet = workbook.add_worksheet('DT')
	ln_worksheet = workbook.add_worksheet('LN')
	na_worksheet = workbook.add_worksheet('NA')
	dropped_worksheet = workbook.add_worksheet('Dropped')
	excel_jobs()
	workbook.close()
	print("\nProcess Completed!")
	a=input("Press Enter to exit")
	if a:
		exit()
		
elif proceed == "2":
	print("Quitting Process...")
else:
	print("Invalid Input! Quitting the process")



